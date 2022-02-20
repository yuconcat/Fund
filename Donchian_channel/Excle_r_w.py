# -*- coding = utf-8 -*-
# @Time : 2021/1/25 21:01
# @Author : xxx
# @File : Excle_r_w.py
# @Software : PyCharm

import openpyxl
from construction import quick_sort

# 获得第n列数据
def get_column_values(path, n):
    column_list = []
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row  # 获得行数
    for i in range(0, rows):
        a = table.cell(row=i + 1, column=n).value
        column_list.append(a)
    excel.save(path)
    return column_list


# 获得第n行数据
def get_rows_values(path, n):
    rows_list = []
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    cols = table.max_column  # 获得列数
    for i in range(0, cols):
        a = table.cell(row=n, column=i + 1).value
        rows_list.append(a)
    excel.save(path)
    return rows_list


# 在第n列写入数据
def write_column_to_excle(path, n, data_list, column_name):
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row
    # 在第一行写入列名
    table.cell(row=1, column=n).value = column_name
    # 在第二行开始写入
    for i in range(1, rows):
        if i <= len(data_list):
            table.cell(row=i + 1, column=n).value = data_list[i - 1]
        elif i > len(data_list):
            break

    excel.save(path)
    #print("写入成功")

#获得列数
def get_columns_number(path):
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row
    excel.save(path)
    return rows


def get_all_data(path):
    list = []
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row  # 获得行数
    cols = table.max_column  # 获得列数
    for i in range(rows):
        list.append([])
        for j in range(cols):
            num = table.cell(row=i + 1, column=j + 1).value
            list[i].append(num)
    excel.save(path)
    return list

#list[0][i]是时间，list[1][i]是其他数据
def w_col_to_ex_by_time(path, n, list, column_name):
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row
    # 在第一行写入列名
    table.cell(row=1, column=n).value = column_name
    # 在第二行开始写入
    for j in range(0, len(list)):
        for i in range(1, rows):
            if (list[j][0] == table.cell(row=i + 1, column=1).value):
                table.cell(row=i + 1, column=n).value = list[j][1]
                continue

    excel.save(path)
    #print("写入成功")


# 获得第n列的最大值数据
def get_column_max_values(path, n):
    column_list = []
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row  # 获得行数
    for i in range(0, rows):
        a = table.cell(row=i + 1, column=n).value
        if a!=None:
            column_list.append(a)
    excel.save(path)
    list_max = quick_sort.Quicksort_get_list(column_list[1:])
    return list_max[-1]




# 获得第n列的最小值数据
def get_column_min_values(path, n):
    column_list = []
    excel = openpyxl.load_workbook(path)
    table = excel.get_sheet_by_name('Sheet')
    rows = table.max_row  # 获得行数
    for i in range(0, rows):
        a = table.cell(row=i + 1, column=n).value
        if a!=None:
            column_list.append(a)
    list_min = quick_sort.Quicksort_get_list(column_list[1:])

    excel.save(path)
    return list_min[0]

if __name__ == '__main__':
    pass
    Name = "招商中证白酒指数(LOF)"  # input("输入基金的名称:")
    fundcode = "161725"  # input("输入基金的代码:")
    path = "D:\\" + "基金\\" + Name + fundcode + "\\" + Name + ".xlsx"
    print(get_column_max_values(path, 3))

    print(get_column_min_values(path, 3))
