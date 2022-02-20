# -*- coding = utf-8 -*-
# @Time : 2021/6/27 19:35
# @Author : xxx
# @File : fund_name_excle_to_mysql.py
# @Software : PyCharm


import pymysql
from openpyxl.reader.excel import load_workbook

def importExcelToMysql(cur, path,tablename):
    num = 1
    # 读取excel文件
    workbook = load_workbook(path)
    # 获得所有工作表的名字
    sheets = workbook.sheetnames
    # 获得第一张表
    worksheet = workbook[sheets[0]]

    # 将表中每一行数据读到sqlstr数组中
    for row in worksheet.rows:
        if num == 1:
            num += 1
            continue
        sqlstr = []
        for cell in row:
            sqlstr.append(cell.value)

        valuestr = [int(sqlstr[0]), str(sqlstr[1]), float(sqlstr[2]), float(sqlstr[3]),
                    str(sqlstr[4]), float(sqlstr[5]), str(sqlstr[6]), str(sqlstr[7])]
        # valuestr = tuple(valuestr)
        # 将每行数据存到数据库中
        cur.execute(f"insert into {tablename}"
                    f"(fundcode, fund_name, ZRDWJZ, JRDWJZ,day_value,day_value_rate,subscription_status,service_charge)"
                    " VALUES(%s, %s, %s, %s, %s, %s, %s, %s);", valuestr)

def connect_db(dbname,tablename,exclepath):
    # 与数据库建立连接
    conn = pymysql.connect(host='127.0.0.1', user='root', password='981128',
                           database=dbname, port=3306, charset='utf8')

    # 创建游标链接
    cur = conn.cursor()

    #创建表
    # 如果存在students这个表则删除
    sql=f"drop table if exists {tablename};"
    cur.execute(sql)
    # 创建表
    sql = f"CREATE TABLE {tablename}" \
          f"(fundcode int(10)," \
          f"fund_name varchar(20)," \
          f"ZRDWJZ decimal(6,4)," \
          f"JRDWJZ decimal(6,4)," \
          f"day_value decimal(6,4)," \
          f"day_value_rate decimal(6,4)," \
          f"subscription_status varchar(5)," \
          f"service_charge varchar(5)," \
          f"PRIMARY KEY (fundcode));"
    cur.execute(sql)

    # 将excel中的数据导入数据库中
    importExcelToMysql(cur, exclepath,tablename)
    # 关闭游标链接
    cur.close()
    conn.commit()
    # 关闭数据库服务连接, 释放内存
    conn.close()


def create_db(dbname):
    # 与数据库建立连接
    conn = pymysql.connect(host='127.0.0.1', user='root', password='981128',
                           port=3306, charset='utf8')
    # 创建游标链接
    cur = conn.cursor()

    # 新建一个database
    sql = f"drop database if exists {dbname};"
    cur.execute(sql)
    sql =f"create database {dbname};"
    cur.execute(sql)
    sql = f"use {dbname}"
    cur.execute(sql)
    # 关闭游标链接
    cur.close()
    conn.commit()
    # 关闭数据库服务连接, 释放内存
    conn.close()

if __name__ == '__main__':
    dbname= "jijin"
    Name = "全部基金"
    tablename = "fundname"
    exclepath = "D:\\基金\\" + Name  +".xlsx"
    connect_db(dbname,tablename,exclepath)
    print("导入sql成功")

