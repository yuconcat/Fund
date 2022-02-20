# -*- coding = utf-8 -*-
# @Time : 2021/6/25 17:21
# @Author : xxx
# @File : detail_data_excle_to_mysql.py
# @Software : PyCharm


import pymysql
from openpyxl.reader.excel import load_workbook

def importExcelToMysql(cur, path,tablename,fundcode):
    num = 1
    # 读取excel文件
    workbook = load_workbook(path)
    # 获得所有工作表的名字
    sheets = workbook.sheetnames
    # 获得第一张表
    worksheet = workbook[sheets[0]]

    fundcode = int(fundcode, base=10)  #把fundcode转换成数字

    # 将表中每一行数据读到sqlstr数组中
    for row in worksheet.rows:
        if num == 1:
            num += 1
            continue
        sqlstr = []
        for cell in row:
            sqlstr.append(cell.value)

        valuestr = [int(fundcode),str(sqlstr[0]), float(sqlstr[1]), float(sqlstr[2]), float(sqlstr[3])]
        # valuestr = tuple(valuestr)
        # 将每行数据存到数据库中

        cur.execute(f"insert into {tablename}(fundcode,Time, DWJZ, LJJZ, RZZL)"
                    " VALUES(%s ,%s, %s, %s, %s);",valuestr)


def connect_db(dbname,tablename,exclepath,fundcode):
    # 与数据库建立连接
    conn = pymysql.connect(host='127.0.0.1', user='root', password='981128',
                           database=dbname, port=3306, charset='utf8')

    # 创建游标链接
    cur = conn.cursor()

    #创建表
    # 如果存在students这个表则删除
    # sql=f"drop table if exists {tablename};"
    # cur.execute(sql)
    # 创建表
    sql = f"CREATE TABLE if not exists {tablename}(id int(8) not null auto_increment,fundcode int(10)," \
          f"Time date,DWJZ decimal(8,4),LJJZ decimal(8,4),RZZL decimal(8,4),PRIMARY KEY (id));"
    cur.execute(sql)


    # 将excel中的数据导入数据库中
    importExcelToMysql(cur, exclepath,tablename,fundcode)
    # 关闭游标链接
    cur.close()
    conn.commit()
    # 关闭数据库服务连接, 释放内存
    conn.close()
    print("导入sql成功")


def create_db(dbname):
    # 与数据库建立连接
    conn = pymysql.connect(host='127.0.0.1', user='root', password='981128',
                           port=3306, charset='utf8')
    # 创建游标链接
    cur = conn.cursor()

    # 新建一个database
    sql = f"drop database if exists {dbname};"
    cur.execute(sql)
    sql =f"create database {dbname};"
    cur.execute(sql)
    sql = f"use {dbname}"
    cur.execute(sql)
    # 关闭游标链接
    cur.close()
    conn.commit()
    # 关闭数据库服务连接, 释放内存
    conn.close()


if __name__ == '__main__':
    dbname= "jijin"    #不动
    Name = "招商中证白酒指数(LOF)"   #改
    fundcode = "161725"   #改
    tablename = "fund"    #不动
    exclepath = "D:\\基金\\" + Name + fundcode +"\\" + Name +".xlsx"   #不动
    exclepath_sql = "D:\\基金\\" + Name + fundcode + "\\" + Name + ".xlsx"
    connect_db(dbname,tablename,exclepath,fundcode)   #不动


