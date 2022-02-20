# -*- coding = utf-8 -*-
# @Time : 2021/1/23 21:03
# @Author : xxx
# @File : leadin_in_excle.py
# @Software : PyCharm

import openpyxl
import json
import xlwt


# 导入xls文件
def to_excle(Name, column_name, path):
    #组合路径
    r_flieName0 = "\\" + Name + "_" + column_name[0] + "_values.json"
    r_flieName1 = "\\" + Name + "_" + column_name[1] + "_values.json"
    r_flieName2 = "\\" + Name + "_" + column_name[2] + "_values.json"
    r_flieName3 = "\\" + Name + "_" + column_name[3] + "_values.json"
    #打开文件
    read_fp0 = open(path + r_flieName0, 'rb')
    read_fp1 = open(path + r_flieName1, 'rb')
    read_fp2 = open(path + r_flieName2, 'rb')
    read_fp3 = open(path + r_flieName3, 'rb')
    #加载数据
    fileJson0 = json.load(read_fp0)
    fileJson1 = json.load(read_fp1)
    fileJson2 = json.load(read_fp2)
    fileJson3 = json.load(read_fp3)

    # 创建一个工作簿
    f = xlwt.Workbook(encoding="utf-8", style_compression=0)
    # 创建一个sheet
    sheet1 = f.add_sheet("sheet1", cell_overwrite_ok=True)

    # 写内容,row行数，column列数（都是从0开始）
    for i in range(0, len(column_name)):
        sheet1.write(0, i, column_name[i])  # 填入列名

    for i in range(0, len(fileJson0)):
        sheet1.write(i + 1, 0, fileJson0[i])

    for i in range(0, len(fileJson1)):
        a = float(fileJson1[i])
        sheet1.write(i + 1, 1, a)

    for i in range(0, len(fileJson2)):
        a = float(fileJson2[i])
        sheet1.write(i + 1, 2, a)

    for i in range(0, len(fileJson3)):
        a = fileJson3[i]
        aa = float(a)  # 里面含有空元素，无法转换
        sheet1.write(i + 1, 3, aa)

    # for i in range(0, len(fileJson3)):
    #     sheet1.write(i + 1, 3, float(fileJson3[i]))   #不能直接在write里把字符强制转换成float

    # for j in range(0,len(row)):
    #     sheet1.write(i+1,j,data[j])     #数据

    f.save(path + "\\" + Name + '.xls')  # 保存数据表
    read_fp0.close()
    read_fp1.close()
    read_fp2.close()
    read_fp3.close()
    print(Name + "导入excle成功")


# 导入xlsx文件
def to_excle2(Name, column_name, path):
    r_flieName0 = "\\" + Name + "_" + column_name[0] + "_values.json"
    r_flieName1 = "\\" + Name + "_" + column_name[1] + "_values.json"
    r_flieName2 = "\\" + Name + "_" + column_name[2] + "_values.json"
    r_flieName3 = "\\" + Name + "_" + column_name[3] + "_values.json"
    read_fp0 = open(path + r_flieName0, 'rb')
    read_fp1 = open(path + r_flieName1, 'rb')
    read_fp2 = open(path + r_flieName2, 'rb')
    read_fp3 = open(path + r_flieName3, 'rb')
    fileJson0 = json.load(read_fp0)
    fileJson1 = json.load(read_fp1)
    fileJson2 = json.load(read_fp2)
    fileJson3 = json.load(read_fp3)

    wookbook = openpyxl.Workbook()
    sheet = wookbook.active

    for i in range(0, len(column_name)):
        sheet.cell(row=1, column=i + 1).value = column_name[i]  # 填入列名

    for i in range(0, len(fileJson0)):
        sheet.cell(row=i + 2, column=1).value = fileJson0[i]  # 填入的是日期-->字符串

    for i in range(0, len(fileJson1)):
        a = float(fileJson1[i])
        sheet.cell(row=i + 2, column=2).value = a

    for i in range(0, len(fileJson2)):
        if (len(fileJson2[i])==0):
            a=0
        else:
            a = float(fileJson2[i])
        sheet.cell(row=i + 2, column=3).value = a

    for i in range(0, len(fileJson3)):
        a = float(fileJson3[i])
        sheet.cell(row=i + 2, column=4).value = a

    sheet.cell(row=len(fileJson3)+2, column=4).value = 0

    wookbook.save(path + "\\" + Name + '.xlsx')
    read_fp0.close()
    read_fp1.close()
    read_fp2.close()
    read_fp3.close()
    print(Name + "导入excle成功")
